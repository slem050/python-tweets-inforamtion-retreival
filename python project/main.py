from openpyxl import load_workbook
import nltk
from nltk.corpus import stopwords
from string import punctuation
from nltk.stem import PorterStemmer
import collections
from textblob import TextBlob

print('converting excel file to csv \n')
f= open("text.txt","w+")
wb = load_workbook(filename='data.xlsx', read_only=True)
ws = wb['sheet']
iterRows=iter(ws.rows)
next(iterRows)
for row in iterRows:
    if('Brasil' in str(row[5].value) or 'brasil' in str(row[5].value)):
        text=str(row[6].value).split(' ')
        for word in text:
            f.write(word+' ')
        f.write('\n')
f.close()
wb.close()
print('prepearing tokenization \n')
f=open("text.txt","r")
tW= open("Tokens.txt","w+")
tokens=[]
for line in f:
    tLine=repr(nltk.word_tokenize(line))
    tW.write(tLine+'\n')
f.close()
tW.close()
print('case foldin text... \n')
stopWords=open("Tokens.txt");
appendFile = open('caseFolding.txt', 'w+')
for words in stopWords:
    words = eval(words)
    for r in words:
        appendFile.write(r.lower()+' ')
    appendFile.write('\n')
appendFile.close()
print('stop words removal\n')
stop_words = set(stopwords.words('english')+list(punctuation))
Tokens = open("caseFolding.txt")
appendFile = open('stopWordsRemoval.txt', 'w+')
for line in Tokens:
    words=line.split(' ')
    lis=list()
    for r in words:
        if not r in stop_words and r!='\n':
            lis.append(r)
    appendFile.write(str(lis)+'\n')
appendFile.close()
print('stemming text\n')
o=open("stopWordsRemoval.txt");
appendFile = open('stemming.txt', 'w+')
ps = PorterStemmer()
for line in o:
    line=eval(line)
    for r in line:
        word=r;
        word=ps.stem(word)
        appendFile.write(word+' ')
    appendFile.write('\n')
appendFile.close()
with open('stemming.txt') as f:
    c = collections.Counter(f.read().split())
appendFile = open('dictionary.txt', 'w+')
for word in c:
    appendFile.write(word+' '+str(c[word])+'\n')
appendFile.close()
text=open('stemming.txt','r')
appendFile = open('emotionsResults.txt', 'w+')
for line in text:
    blob=TextBlob(line)
    x=blob.sentiment.polarity;
    state='nectural'
    if x!=0:
        if(x>0):
            state='positive'
        else:
            state='negative'
    appendFile.write(state+"-"+line)
text.close()
appendFile.close()
AppendFile= open("datesAndGeoAndFeelings.txt","w+")
wb = load_workbook(filename='data.xlsx', read_only=True)
ws = wb['sheet']
iterRows=iter(ws.rows)
next(iterRows)
emotions=open('emotionsResults.txt','r')
for f, b in zip(iterRows, emotions):
    text=str(f[0].value)+','+str(f[5].value)+','+b.split('-')[0]
    AppendFile.write(text+'\n')
AppendFile.close()
wb.close()
AppendFile= open("FeelingsOnDailyBasis.txt","w+")
emotions=open('datesAndGeoAndFeelings.txt','r')
currentDate=''
counN=0
counP=0
countK=0
for line in emotions:
    l=line.split(',')
    if currentDate!=l[0]:
        if currentDate!='':
            AppendFile.write(currentDate+' postive tweets: '+str(counP)+' negative tweets '+str(counN)+' nectural: '+str(countK)+'\n')
            print(currentDate+' postive tweets: '+str(counP)+' negative tweets '+str(counN)+' nectural: '+str(countK)+'\n')
        counN = 0
        counP = 0
        countK = 0
        currentDate=l[0]
    else:
        if l[2].rstrip()=='negative':
            counN=counN+1
        if l[2].rstrip()=='positive':
            counP=counP+1
        if l[2].rstrip()=='nectural':
            countK=countK+1

AppendFile.close()
emotions.close()
appendFile=open('TweetsPerGeo.txt','w+')
o=open('datesAndGeoAndFeelings.txt','r')
list={}
pTweets={}
nTweets={}
neTweets={}
for line in o:
    l=line.split(',')
    if l[1] in list:
        list[l[1]]=list[l[1]]+1
    else:
        list[l[1]]=1
    if l[2].rstrip()== 'positive':
        if l[1] in pTweets:
            pTweets[l[1]] = pTweets[l[1]] + 1
        else:
           pTweets[l[1]] = 1
    if l[2].rstrip()=='negative':
        if l[1] in nTweets:
            nTweets[l[1]] = nTweets[l[1]] + 1
        else:
           nTweets[l[1]] = 1
    if l[2].rstrip()=='nectural':
        if l[1] in neTweets:
            neTweets[l[1]] = neTweets[l[1]] + 1
        else:
           neTweets[l[1]] = 1

list={k: v for k, v in sorted(list.items(), key=lambda item: item[1])}
pTweets={k: v for k, v in sorted(pTweets.items(), key=lambda item: item[1])}
nTweets={k: v for k, v in sorted(nTweets.items(), key=lambda item: item[1])}
neTweets={k: v for k, v in sorted(neTweets.items(), key=lambda item: item[1])}

for g in list:
    appendFile.write(g+": "+str(list[g])+'\n')
print('positive and negative and nectural tweets per each geo:')
print('all tweets count: '+str(list)+'\nnectural: '+str(neTweets)+'\npostive: '+str(pTweets)+'\nnegative:'+str(nTweets))





